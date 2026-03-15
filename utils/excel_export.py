import pandas as pd
import io
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side, Fill, PatternFill
from openpyxl.utils import get_column_letter

def export_stock_statement_to_excel(stocks, month_name, year):
    output = io.BytesIO()
    
    # Define Columns exactly as per template
    headers = [
        "PRODUCT DESCRIPTION",
        "OPENING STOCK",
        "RECEIVE",
        "SALE RETURN QUANTITY",
        "REPLACE + OTHERS",
        "TOTAL QTY",
        "SALE QUANTITY",
        "P/R QUANTITY",
        "REPLACE + OTHERS",
        "CLOSING STOCK"
    ]
    
    data = []
    total_qty = 0
    total_value = 0 # Placeholder if value logic is needed later
    
    for s in stocks:
        row = [
            s.product_ref.product_name if s.product_ref else "NA",
            s.opening_stock,
            s.received_stock,
            s.sale_return_qty,
            s.replace_others_in,
            s.total_quantity,
            s.sales,
            s.pr_quantity,
            s.replace_others_out,
            s.closing_stock
        ]
        data.append(row)
        total_qty += s.closing_stock
        # total_value += ... (could sum based on rate if available)

    df = pd.DataFrame(data, columns=headers)
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=2, sheet_name='Stock Statement')
        workbook = writer.book
        worksheet = writer.sheets['Stock Statement']
        
        # 1. Main Titles
        title1 = "FEMINA LIFESCIENCES PVT LTD"
        title2 = f"STOCK & SALES STATEMENT [{month_name.upper()} {str(year)[-2:]}]"
        
        worksheet.merge_cells('A1:J1')
        worksheet.merge_cells('A2:J2')
        
        cell1 = worksheet['A1']
        cell2 = worksheet['A2']
        
        cell1.value = title1
        cell2.value = title2
        
        # Styling Titles
        title_font = Font(name='Arial', size=14, bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        
        cell1.font = title_font
        cell1.alignment = center_align
        cell2.font = Font(name='Arial', size=12, bold=True)
        cell2.alignment = center_align
        
        # 2. Header Styling
        header_font = Font(bold=True, color="000000")
        header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col_num in range(1, 11):
            cell = worksheet.cell(row=3, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
            
        # 3. Data Styling & Borders
        for row in range(4, 4 + len(data)):
            for col in range(1, 11):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border
                if col > 1: # Numeric columns
                    cell.alignment = Alignment(horizontal='right')
        
        # 4. Footer (Total Quantity & Total Value)
        footer_row = 4 + len(data)
        worksheet.cell(row=footer_row, column=1, value="TOTAL QUANTITY").font = Font(bold=True)
        worksheet.cell(row=footer_row, column=10, value=total_qty).font = Font(bold=True)
        worksheet.cell(row=footer_row, column=10).alignment = Alignment(horizontal='right')
        
        worksheet.cell(row=footer_row + 1, column=1, value="TOTAL VALUE").font = Font(bold=True)
        # worksheet.cell(row=footer_row + 1, column=10, value=total_value).font = Font(bold=True)
        
        # 5. Column Widths
        worksheet.column_dimensions['A'].width = 40
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            worksheet.column_dimensions[col].width = 15

    output.seek(0)
    return output
